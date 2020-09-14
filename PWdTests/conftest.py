from datetime import datetime

import pytest
import openpyxl
import pytest_html

from POMDoc.BasePage import BasePage
from WdTests.DriverFactory import getDriver
from config import logger as log
from py._xmlgen import html

def pytest_addoption(parser):
    parser.addoption(
        "--browser_name", action="store", default="chrome"
    )

@pytest.fixture(scope='class')
def setupOhrm(request):
    brName = request.config.getoption('--browser_name')
    # driver = getDriver('chrome')
    driver = getDriver(brName)
    driver.set_page_load_timeout(20)
    driver.implicitly_wait(20)
    driver.maximize_window()
    driver.get('https://qtpsudhakarblog-trials65141.orangehrmlive.com/')
    print("Open Orange hrm application")
    log.info('application opened from setup')
    request.cls.driver = driver
    print(request.node.name)
    yield
    driver.close()
    print("Closed Application")

@pytest.fixture(scope='class',params=['chrome','firefox'])
def multisetupOhrm(request):
    driver = getDriver(request.param)
    driver.set_page_load_timeout(20)
    driver.implicitly_wait(20)
    driver.maximize_window()
    driver.get('https://qtpsudhakarblog-trials65141.orangehrmlive.com/')
    print("Open Orange hrm application")
    request.cls.driver = driver
    yield
    driver.close()
    print("Closed Application")

@pytest.fixture(scope='class')
def setupOhrmPOM():
    Ohrm = BasePage()
    Ohrm.OpenApplication('chrome')
    yield
    Ohrm.CloseApplication()

@pytest.fixture(scope='class')
def getTCData(request):
    flPath = r"C:\Users\envy\Desktop\PythonForTestersProject\Data\Demo.xlsx"
    wb = openpyxl.load_workbook(flPath)
    sht = wb['Sheet2']
    rc = sht.max_row
    cc = sht.max_column
    dc = {}
    for r in range(2, rc + 1):
        tcName = sht.cell(r, 1).value
        if tcName == request.node.name:
            for c in range(1, cc + 1):
                kname = sht.cell(1, c).value
                kval = sht.cell(r, c).value
                dc.update({kname: kval})
            break
    wb.close()
    request.cls.data = dc

@pytest.mark.optionalhook
def pytest_html_results_table_header(cells):
    cells.insert(2, html.th('Description'))
    cells.insert(3, html.th('Time', class_='sortable time', col='time'))
    cells.pop()

@pytest.mark.optionalhook
def pytest_html_results_table_row(report, cells):
    cells.insert(2, html.td(report.description))
    cells.insert(3, html.td(datetime.utcnow(), class_='col-time'))
    cells.pop()

@pytest.mark.hookwrapper
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()
    report.description = str(item.function.__doc__)
    extra = getattr(report,'extra',[])
    # file_name = item.__name__+'_'+item.function.__name__+'/'+'.png'
    if report.when == 'call' or report.when == "setup":
        xfail = hasattr(report, 'wasxfail')
        if (report.skipped and xfail) or (report.failed and not xfail):
            # SeleniumTests/test_ohrm.py::Test_demo::test_login
            bflname = report.nodeid
            file_name = bflname.replace("/",'_').replace('.py','_').replace('::','_')+'.png'
            # file_name = report.nodeid.replace("::", "_")+".png"

            if file_name:
                html = '<div><img src="data:image/gif;base64,%s" alt="screenshot" style="width:304px;height:228px;" ' \
                       'onclick="window.open(this.src)" align="right"/></div>' % BasePage.getScreenshotBase()
                # html = '<div><img src="%s" alt="screenshot" style="width:304px;height:228px;" ' \
                #        'onclick="window.open(this.src)" align="right"/></div>' % BasePage.getScreenshot(file_name)

                extra.append(pytest_html.extras.html(html))
        report.extra = extra
