import openpyxl

flPath = r"C:\Users\envy\Desktop\PythonForTestersProject\Data\Demo.xlsx"
wb = openpyxl.load_workbook(flPath)
# sht = wb['Sheet1']
sht = wb.active
rc = sht.max_row
cc = sht.max_column
for r in range(1,rc+1):

    for c in range(1,cc+1):
        print(sht.cell(r, c).value)
wb.close()