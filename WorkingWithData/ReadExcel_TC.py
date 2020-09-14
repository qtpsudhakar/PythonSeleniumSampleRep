import openpyxl

flPath = r"C:\Users\envy\Desktop\PythonForTestersProject\Data\Demo.xlsx"
wb = openpyxl.load_workbook(flPath)
sht = wb['Sheet2']
# sht = wb.active
rc = sht.max_row
cc = sht.max_column

for r in range(2,rc+1):
    tcName = sht.cell(r, 1).value
    global dc
    dc = {}
    if tcName=='TC_Demo_002':
        for c in range(1,cc+1):
            kname = sht.cell(1, c).value
            kval = sht.cell(r, c).value

            dc.update({kname:kval})
        break
wb.close()
print(dc)