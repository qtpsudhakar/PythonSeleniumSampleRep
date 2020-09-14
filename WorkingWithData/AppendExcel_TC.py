import openpyxl

flPath = r"C:\Users\envy\Desktop\PythonForTestersProject\Data\Demo.xlsx"
wb = openpyxl.load_workbook(flPath)
# sht = wb['Sheet1']
sht = wb.active
rc = sht.max_row
cc = sht.max_column

sht.cell(rc+1, 1).value = "new value"
wb.save(flPath)
wb.close()