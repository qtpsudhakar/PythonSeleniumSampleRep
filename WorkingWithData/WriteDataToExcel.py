from openpyxl import Workbook

wb = Workbook()
sht = wb.create_sheet('demo')
sht.cell(1,1,"empno")
sht.cell(1,2,"empname")

wb.save(r"C:\Users\envy\Desktop\PythonForTestersProject\Data\Demo1.xlsx")
wb.close()